from bs4 import BeautifulSoup 
import requests
import time
import pandas as pd
from openpyxl import load_workbook
import os

# User-Agent, kendimizi gerçek bir kullanıcı olarak göstermek için kullanıcı bilgilerimiz
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
}


def convert_storage_to_gb(storage_string):
    try:
        if "TB" in storage_string:
            # "TB" kısmını kaldırıyoruz ve sonucu float'a çeviriyoruz
            tb_value = float(storage_string.replace("TB", "").strip())
            # TB'yi GB'ye çeviriyoruz
            gb_value = tb_value * 1024
            # Sonucu integer'a çeviriyoruz
            return int(gb_value)
        else:
            # Eğer "TB" yoksa, stringi olduğu gibi float'a çevirip döndür
            # Örneğin, 500GB ise sadece sonundaki GB'yi kaldırıp integer'a çevirelim
            gb_value = float(storage_string.replace("GB", "").strip())
            return int(gb_value)
    except ValueError:
        # Hata durumunda (örneğin, geçersiz bir format) 0 döndürüyoruz
        return 0
    

def append_to_excel(df, sheet_name='Sheet1'):
        
        #excel dosyasının yolu
        file_path = './output.xlsx'
        
        #fonksiyona soktuğumuz dataframe i pandas kütüphanesindeki dataframe tipine eşitle
        df_results = pd.DataFrame(df)

        try:
            #pandas kullanarak bir writer objesi oluştur, bu obje bizim excelde yazı yazmamızı sağlayan başka bir yazılım olarak düşünebilirsiniz
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                
                book = writer.book
                if sheet_name in book.sheetnames:
                    startrow = book[sheet_name].max_row

                else:
                    startrow = 0
                df_results.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=startrow == 0)
        except Exception as e:
                print(f"Error loading workbook: {e}")

        print(f"Metrikler {file_path} yoluna kaydedildi.")



    
# maxPageNumber
maxPageNumber = 25
productNumber = 0


# 1 den başlayarak maksimum sayfa sayısı + 1 kadar bu döngüyü uygula
for i in range(1,maxPageNumber+1):
    
    pageUrl = "https://www.laptopsdirect.co.uk/ct/laptops-and-netbooks/laptops?pageNumber=" + str(i)
    print("Going to " + pageUrl)
    
    # scrapleyeceğimiz sayfaya bağlanıyoruz
    pageToScrapLinks = requests.get(pageUrl, headers=headers)

    #2 saniye bekliyoruz
    time.sleep(2)

    #eğer bağlantı başarılıysa yani karşıdan 200 cevabını alıyorsak devam et
    if pageToScrapLinks.status_code == 200:

        #sayfanın html kodunu al ve bunu soup değişkenine eşitle
        soup = BeautifulSoup(pageToScrapLinks.content, "html.parser") #html content
        print("Page successfully scraped!")

        #class ı OfferBox border-radius... olan bütün divleri bir listeye al
        divOfferBox = soup.find_all('div', class_='OfferBox border-radius-large border-style-solid border-width-1 border-neutral-300 margin-y-1 b-row margin-x-0 position-relative')
        
        #divOfferBox listesinin içinde bulunan her bir eleman yani item sayısı kadar bu döngüyü çalıştır
        for item in divOfferBox:
            
            #FİYAT BİLGİSİ -------------------------------------------------------------------------------------------------------------
            #divOfferBox listesinin içindeki mevcut itemin içinde class ı OfferBoxPrice b-col-4... olan elementi bul
            offerBoxPrice = item.find('div', class_='OfferBoxPrice b-col-4 b-col-xl-3 padding-y-1 padding-left-0')

            #offerBoxPrice div inin içinde olan class ı offerprice olan span elementini bul
            offerPriceSpan = offerBoxPrice.find('span',class_='offerprice')

            #offerPriceSpan ın içindeki spanları bul ve bir priceSpans listesine kaydet
            priceSpans = offerPriceSpan.find_all('span')

            #priceSpans listesinin ilk elemanını al ve bunu rawPriceDataya eşitle FİYAT bilgisi bu
            rawPriceData = priceSpans[0].text
            #-----------------------------------------------------------------------------------------------------------------------
               
            
            #TEKNİK ÖZELLİKLER---------------------------------------------------------------------------------------------------------
            #divOfferBox listesinin içindeki mevcut itemin içinde class ı OfferBoxProdInfo b-col-5... olan div i bul
            offerprodinfo = item.find('div',class_='OfferBoxProdInfo b-col-5 b-col-xl-6 padding-1')
            
            #offerProdinfo divinin içinde olan classı productInfo olan div i bul
            prodinfo = offerprodinfo.find('div',class_='productInfo')

            #prodinfo divinin içindeki ul elementini bul
            ul = prodinfo.find('ul')
            

            #ul elementinin içinde class ı proddesctitle olan span elemanlarını bul ve bir listeye kaydet
            proddesctitles = ul.find_all('span',class_='proddesctitle')


            try:
                #proddestitles listesinin içindeki 2. elementi bul
                gpucheck = proddesctitles[1].text
                
                #ul elementinin içinde class ı proddescvalue olan spanları bul ve bir listeye kaydet
                productvalue = ul.find_all('span',class_='proddescvalue')
                
                #eğer gpucheck değişkeni "Graphics card" yazısına eşitse
                if gpucheck == "Graphics card":
                    print("************************** WITH GPU *********************")

                    try:
                        #productvalue listesinin içindeki ilk eleman yani İŞLEMCİ
                        processor = productvalue[0].text

                        #productvalue listesinin içindeki ilk eleman yani GPU
                        gpu = productvalue[1].text

                        #productvalue listesinin içindeki ilk eleman yani RAM
                        rawRamData = productvalue[3].text

                        #productvalue listesinin içindeki ilk eleman yani SSD
                        rawSsdData = productvalue[4].text


                        #ram stringinin arasındaki boşluğu ve GB yazısını kaldır ve bunu Stringden integera dönüştür
                        ram = rawRamData.replace(" ", "")
                        ram = ram.replace("GB", "").strip()
                        ram = int(ram)
     
                        #ssd stringinin arasındaki boşluğu ve GB yazısını kaldır
                        #convertstorage fonksiyonuyla eğer çektiğimiz veride TB varsa bunu 1024 ile çarp ve integer olarak bu değeri döndür
                        ssd = rawSsdData.replace(" ", "")
                        ssd = ssd.replace("GB", "").strip()
                        ssd = convert_storage_to_gb(ssd)
                            
                        #fiyat bilgisinden euro işaretini kaldır ve stringden float veri tipine dönüştür
                        price = rawPriceData.replace("£", "").strip()
                        price = float(price)

                        #bu döngü her çalıştığında productNumber a 1 ekle
                        productNumber = productNumber + 1

                        print("Product Number: " + str(productNumber))
                        print(price)
                        print(processor)
                        print(gpu)
                        print(ram)
                        print(ssd)

                        # Değişkenleri bir veri çerçevesine (DataFrame) dönüştürün
                        data = {
                                'ProductID': [productNumber],
                                'Price': [price],
                                'Processor': [processor],
                                'GPU': [gpu],
                                'RAM': [ram],
                                'SSD': [ssd]
                            }

                        df = pd.DataFrame(data)

                        # DataFrame'i bir Excel dosyasına kaydet
                        append_to_excel(df)

                       



                    except IndexError:
                        print("bad html productvalue is empty")

                #eğer gpucheck değişkeni "Graphics card" yazısına eşit DEĞİLSE
                else:
                    try:
                        print("************************** NO GPU *********************")
                
                        #productvalue listesinin içindeki ilk eleman yani İŞLEMCİ
                        processor = productvalue[0].text

                        #GPU YOK
                        gpu = "-"

                        #productvalue listesinin içindeki ilk eleman yani RAM
                        rawRamData = productvalue[2].text

                        #productvalue listesinin içindeki ilk eleman yani SSD
                        rawSsdData = productvalue[3].text

                        #ram stringinin arasındaki boşluğu ve GB yazısını kaldır ve bunu Stringden integera dönüştür
                        ram = rawRamData.replace(" ", "")
                        ram = ram.replace("GB", "").strip()
                        ram = int(ram)
                        
                        #ssd stringinin arasındaki boşluğu ve GB yazısını kaldır
                        #convertstorage fonksiyonuyla eğer çektiğimiz veride TB varsa bunu 1024 ile çarp ve integer olarak bu değeri döndür
                        ssd = rawSsdData.replace(" ", "")
                        ssd = ssd.replace("GB", "").strip()
                        ssd = convert_storage_to_gb(ssd)
 
                        #fiyat bilgisinden euro işaretini kaldır ve stringden float veri tipine dönüştür
                        price = rawPriceData.replace("£", "").strip()
                        price = float(price)
                        
                        #bu döngü her çalıştığında productNumber a 1 ekle
                        productNumber = productNumber + 1

                        print("Product Number: " + str(productNumber))
                        print(price)
                        print(processor)
                        print(gpu)
                        print(ram)
                        print(ssd)


                         # Değişkenleri bir veri çerçevesine (DataFrame) dönüştürün
                        data = {
                                'ProductID': [productNumber],
                                'Price': [price],
                                'Processor': [processor],
                                'GPU': [gpu],
                                'RAM': [ram],
                                'SSD': [ssd]
                            }

                        df = pd.DataFrame(data)

                        # DataFrame'i bir Excel dosyasına kaydedin
                        append_to_excel(df)


                    #productvalue listesi boş olduğu için bu hata mesajını fırlatıyoruz
                    except IndexError:
                        print("bad html productvalue is empty")
            
            #proddesctitles listesi boş olduğu için bu hata mesajını fırlatıyoruz
            except IndexError:
                print("bad html proddesctitles is empty")

    #siteye attığımız istek başarısız site yok veya html yapısı bozuk
    else:
        print("fail to request")


print("**********DONE***************")

